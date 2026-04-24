import openpyxl
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path("D:/git/dev/dbi_patcher")
DICT_PATH = ROOT / "data/dictionary.xlsx"
BLOCKS_CONFIG = ROOT / "data/blocks.json"

def analyze_disambiguation():
    with BLOCKS_CONFIG.open("r", encoding="utf-8") as f:
        blocks = json.load(f)
    
    wb = openpyxl.load_workbook(DICT_PATH, data_only=True)
    ws = wb["Translations"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    # Збираємо всі варіанти рядків з Excel
    all_data = []
    for row in range(2, ws.max_row + 1):
        full_val = str(ws.cell(row, col_map["Original"]).value or "").strip()
        if not full_val: continue
        
        parts = full_val.split(":", 1)
        prefix = parts[0].rstrip() # Префікс ДО двокрапки з пробілами
        label = prefix.strip()     # Чиста назва
        suffix = parts[1].strip() if len(parts) > 1 else ""
        
        all_data.append({
            "row": row,
            "label": label,
            "prefix": prefix,
            "suffix": suffix,
            "full": full_val
        })

    final_mapping = {} # block_id -> {label: row_info}

    for block_id, terms in blocks.items():
        final_mapping[block_id] = {}
        is_info = block_id.endswith("_INFO")
        
        for term in terms:
            matches = [d for d in all_data if d["label"].lower() == term.lower()]
            
            if not matches:
                continue

            target_match = None
            
            # --- ПРАВИЛА РОЗПОДІЛУ ---
            
            # 1. Серийный номер
            if term == "Серийный номер":
                if block_id == "SD_INFO":
                    target_match = next((m for m in matches if "{:s}" in m["suffix"]), None)
                else:
                    target_match = next((m for m in matches if "{:s}" not in m["suffix"]), None)

            # 2. Заряд батареи
            elif term in ["Заряд батареи", "Заряд батареи (сырое значение)"]:
                target_match = next((m for m in matches if "%" in m["suffix"]), None)

            # 3. Возраст батареи
            elif term == "Возраст батареи":
                if block_id == "CHARGE_INFO":
                    # Версія з % та максимальними пробілами
                    target_match = sorted([m for m in matches if "%" in m["suffix"]], 
                                         key=lambda x: len(x["prefix"]), reverse=True)[0]
                else:
                    # Решта для POWER_INFO
                    target_match = next((m for m in matches if "%" not in m["suffix"]), None)

            # 4. Текущая ёмкость (враховуючи різне написання е/ё)
            elif term.lower() in ["текущая емкость", "текущая ёмкость"]:
                if block_id == "CHARGE_INFO":
                    target_match = next((m for m in matches if "%" in m["suffix"]), None)
                elif block_id == "MAX17050_INFO":
                    target_match = next((m for m in matches if "mah" in m["suffix"].lower()), None)
            
            # Стандартна логіка для решти
            if not target_match:
                if is_info:
                    target_match = next((m for m in matches if "{" in m["suffix"]), matches[0])
                else:
                    target_match = next((m for m in matches if not m["suffix"]), matches[0])

            final_mapping[block_id][term] = target_match

    # Виводимо звіт
    print("### ✅ Фінальний розподіл дублікатів по блоках:")
    for b_id, mapped_terms in final_mapping.items():
        if not mapped_terms: continue
        print(f"\n**{b_id}**:")
        for term, info in mapped_terms.items():
            if term in ["Серийный номер", "Заряд батареи", "Заряд батареи (сырое значение)", "Возраст батареи", "Текущая ёмкость", "Текущая емкость"]:
                print(f"- {term} -> Строка {info['row']}: `{info['full']}`")

if __name__ == "__main__":
    analyze_disambiguation()
