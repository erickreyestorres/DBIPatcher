import openpyxl
import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path("D:/git/dev/dbi_patcher")
DICT_PATH = ROOT / "data/dictionary.xlsx"
BLOCKS_CONFIG = ROOT / "data/blocks.json"

def verify_regex_mapping():
    with BLOCKS_CONFIG.open("r", encoding="utf-8") as f:
        blocks = json.load(f)
    
    wb = openpyxl.load_workbook(DICT_PATH, data_only=True)
    ws = wb["Translations"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    # Збираємо всі значення з Original стовпця
    excel_rows = []
    for row in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row, col_map["Original"]).value or "").strip()
        if cell_val:
            excel_rows.append({"row": row, "val": cell_val})

    all_ok = True
    errors = []

    for block_id, patterns in blocks.items():
        print(f"\nChecking block: {block_id}")
        for pattern in patterns:
            try:
                regex = re.compile(pattern, re.IGNORECASE)
            except re.error as e:
                errors.append(f"Invalid Regex in JSON: {pattern} ({e})")
                all_ok = False
                continue

            matches = []
            for item in excel_rows:
                if regex.search(item["val"]):
                    matches.append(item)
            
            if len(matches) == 0:
                errors.append(f"NOT FOUND: '{pattern}' in block {block_id}")
                all_ok = False
            elif len(matches) > 1:
                match_details = ", ".join([f"row {m['row']} ('{m['val']}')" for m in matches])
                errors.append(f"AMBIGUOUS (found {len(matches)}): '{pattern}' matches: {match_details}")
                all_ok = False
            else:
                # Знайдено рівно один рядок
                pass

    if all_ok:
        print("\n✅ Усі патерни в blocks.json валідні та мають рівно один відповідний рядок у словнику.")
    else:
        print("\n❌ Знайдено проблеми з мапінгом:")
        for err in errors:
            print(f"  - {err}")

if __name__ == "__main__":
    verify_regex_mapping()
