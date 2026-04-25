import openpyxl
import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path("D:/git/dev/dbi_patcher")
DICT_PATH = ROOT / "data/dictionary.xlsx"
BLOCKS_CONFIG = ROOT / "data/blocks.json"

RAW_BLOCKS = {
        "NSP_INSTALL": ["Общий размер передачи", "Общий размер установки", "Место установки", "Не проверять место", "Удалять после установки", "Выключить экран", "Выйти из DBI", "Проверять хеш-сумму NCA", "Патчить требование NNID", "Патчить скриншоты", "Патчить запись видео", "Занулять мастерключ", "Патчить имена (HOS 21+)"],
        "SETTINGS": ["Выключать экран", "Читать дату файлов", "CDIR в листинге", "Форсировать UTF-8", "Использовать MSG_WAITALL в recv", "Запускать точку доступа", "SSID", "Пароль", "Использовать 5 GHz", "Использовать скрытый SSID", "Тема отображения", "Переносить слова", "Выход на рабочий стол", "Журналирование действий", "Подсвечивать файлы обновлений", "Перевернуть экран", "Перевернуть джойконы", "Использовать разгон", "Сохранения только в RO режиме", "Показывать 'Обновить отсюда'", "Папка для бекапа сохранений", "Папка журналов", "Папка для дампа игр", "URL для проверки обновлений", "Показывать прогрев кеша", "Смещать курсор после выделения", "Время гашения экрана в секундах", "Автоповтор кнопок при удержании", "Курсор на обеих панелях", "Показывать часы в статусе", "Показывать секунды", "Интервал автобекапа сейвов, дни", "Показывать размер LFS (долго)", "Проверять хэш при установке", "Создавать LFS папку", "Показывать объединённый NSP", "Добавлять суффикс к имени файла", "Показывать папку 'Mods&Cheats'", "Использовать для неё TitleID", "Android extensions", "Просмотр SD карты", "Просмотр раздела SYSTEM", "Просмотр раздела USER", "Просмотр USB носителей", "Установка через DBIbackend", "Установка с картриджа", "Просмотр сети", "Закладки на SD", "Просмотр установленных игр", "Инструменты", "Просмотр тикетов", "Просмотр сохранений", "Запустить MTP соединение", "Запустить FTP сервер", "Запустить HTTP сервер", "SD Card", "Nand USER", "Nand SYSTEM", "Installed games", "SD Card install", "NAND install", "Saves", "Album", "Gamecard", "Пользовательские хранилища"],
        "FW_INFO": ["Версия прошивки", "Платформа", "Хэш версии", "Показываемая версия", "Показываемое название", "DramId", "Сожжено предохранителей", "Тип SoC", "Тип оборудования", "Предназначение", "DeviceId", "Режим HiZ зарядки", "Режим киоска", "Серийный # прочитанный", "Серийный # угаданный", "Язык", "Регион", "Никнейм консоли", "Родительский PIN"],
        "AMS_INFO": ["Версия атмосферы", "Поколение ключа", "Целевая прошивка", "Git commit hash", "Есть исправление RCM бага", "Экзосфера очищает CAL0", "Запись в CAL0 разрешена", "Emummc включён", "Форсировано вклчение USB 3.0", "Поддерживаемая версия HOS"],
        "SD_INFO": ["Производитель", "OEM ID", "Имя продукта", "Ревизия продукта", "Серийный номер", "Дата производства"],
        "POWER_INFO": ["Заряд батареи", "Заряд батареи (сырое значение)", "Тип источника питания", "Зарядка батареи", "Уровень заряда батареи", "Достаточное питание", "Возраст батареи"],
        "CHARGE_INFO": ["Предел входного тока", "Предел входного тока (буст)", "Предел тока быстрой зарядки", "Предел напряжения заряда батареи", "Конфигурация зарядки", "Режим Hi-Z включён", "Зарядка батареи включена", "Маршрут электропитания", "Температура батареи", "Текущая ёмкость", "Напряжение батареи", "Возраст батареи", "Силовая роль", "Источник питания", "Напряжение источника", "Ток источника", "Быстрая зарядка разрешена", "Контроллер источника получен", "OTG Запрошен"],
        "BATTERY_CONTROLLER_INFO": ["RCOMP0", "TempC0", "FullCap", "FullCapNom", "IavgEmpty", "QrTable00", "QrTable10", "QrTable20", "QrTable30", "Сумма %% заряд-разряд"],
        "MAX17050_INFO": ["Текущий заряд", "Текущая ёмкость", "Полная ёмкость", "Ёмкость с завода", "Текущее напряжение", "Текущий ток", "%% заряд-разряд", "Температура", "Средняя температура"],
        "HW_INFO": ["Bluetooth MAC адрес", "Wireless LAN MAC адрес", "Configuration Id1", "Серийный номер", "Код партії батареї", "Экран"],
        "ACTIVITY_INFO": ["RTC запущены (приблизительно)", "Дата первого игрового события", "Общее число игровых сессий", "Число уникальних зіграних ігор", "Загальний час запущених ігор", "Загальний час активної гри"]
}

def escape_with_exact_spaces(text):
    # Екрануємо все крім пробілів
    res = re.escape(text)
    res = res.replace(r'\ ', ' ')
    
    # Функція для заміни пробілів на точну кількість \s{N}
    def replacer(match):
        spaces = match.group(0)
        if len(spaces) >= 2:
            return f"\\s{{{len(spaces)}}}"
        return spaces

    res = re.sub(r' +', replacer, res)
    return f"^{res}$"

def transform_to_regex():
    wb = openpyxl.load_workbook(DICT_PATH, data_only=True)
    ws = wb["Translations"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    all_data = []
    for row in range(2, ws.max_row + 1):
        full_val = str(ws.cell(row, col_map["Original"]).value or "").strip()
        if not full_val: continue
        parts = full_val.split(":", 1)
        prefix = parts[0].rstrip()
        label = prefix.strip()
        suffix = parts[1].strip() if len(parts) > 1 else ""
        all_data.append({"row": row, "label": label, "prefix": prefix, "suffix": suffix, "full": full_val})

    final_regex_blocks = {}
    for b_id, terms in RAW_BLOCKS.items():
        is_info = b_id.endswith("_INFO")
        regex_list = []
        for term in terms:
            matches = [d for d in all_data if d["label"].lower() == term.lower()]
            if not matches: continue
            
            target = None
            if term == "Серийный номер":
                if b_id == "SD_INFO":
                    target = next((m for m in matches if "{:s}" in m["suffix"]), None)
                else:
                    target = next((m for m in matches if "{:s}" not in m["suffix"]), None)
            elif term in ["Заряд батареи", "Заряд батареи (сырое значение)"]:
                target = next((m for m in matches if "%" in m["suffix"]), None)
            elif term == "Возраст батареи":
                if b_id == "CHARGE_INFO":
                    m_list = [m for m in matches if "%" in m["suffix"]]
                    if m_list:
                        target = sorted(m_list, key=lambda x: len(x["prefix"]), reverse=True)[0]
                else:
                    target = next((m for m in matches if "%" not in m["suffix"]), None)
            elif term == "Текущая ёмкость":
                target = next((m for m in matches if "%" in m["suffix"]), None) if b_id == "CHARGE_INFO" else next((m for m in matches if "mah" in m["suffix"].lower()), None)
            
            if not target:
                if is_info:
                    target = next((m for m in matches if "{" in m["suffix"]), matches[0])
                else:
                    target = next((m for m in matches if not m["suffix"]), matches[0])
            
            regex_list.append(escape_with_exact_spaces(target["full"]))
        if regex_list:
            final_regex_blocks[b_id] = regex_list

    with BLOCKS_CONFIG.open("w", encoding="utf-8") as f:
        json.dump(final_regex_blocks, f, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    transform_to_regex()
