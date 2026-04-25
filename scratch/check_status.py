import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path("D:/git/dev/dbi_patcher")
DICT_PATH = ROOT / "data/dictionary.xlsx"

def check_translation_status():
    wb = openpyxl.load_workbook(DICT_PATH, data_only=True)
    ws = wb["Translations"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    total_rows = ws.max_row - 1
    missing_ua = 0
    missing_en = 0
    
    for row in range(2, ws.max_row + 1):
        orig_val = ws.cell(row, col_map["Original"]).value
        if not orig_val: continue
        
        ua_val = ws.cell(row, col_map["ua"]).value
        en_val = ws.cell(row, col_map["en"]).value
        
        if not ua_val:
            missing_ua += 1
        if not en_val:
            missing_en += 1
            
    print(f"Загальна кількість рядків: {total_rows}")
    print(f"Потребують перекладу (ua): {missing_ua}")
    print(f"Потребують перекладу (en): {missing_en}")

if __name__ == "__main__":
    check_translation_status()
