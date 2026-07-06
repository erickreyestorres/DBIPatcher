#!/usr/bin/env python3
"""Import one translated CSV column into data/dictionary.xlsx.

The CSV files are the easiest format for human review, while dictionary.xlsx is
the source used by the full localization pipeline. This command keeps both
representations synchronized without changing any other language.
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "dictionary.xlsx"

sys.path.insert(0, str(ROOT))

from src.core.text_utils import tokenize  # noqa: E402


def load_translation_csv(path: Path) -> dict[str, str]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["original", "translation"]:
            raise ValueError(
                f"{path}: expected columns original,translation; got {reader.fieldnames}"
            )

        translations: dict[str, str] = {}
        for row_number, row in enumerate(reader, start=2):
            original = tokenize(row.get("original", ""))
            translation = tokenize(row.get("translation", ""))
            if not original or not translation:
                raise ValueError(f"{path}: empty value in row {row_number}")
            if original in translations:
                raise ValueError(f"{path}: duplicate original in row {row_number}")
            translations[original] = translation
    return translations


def bump_workbook_version(workbook) -> str:
    metadata = workbook["Metadata"]
    current = str(metadata["B1"].value or "0.0.0")
    parts = current.split(".")
    parts[-1] = str(int(parts[-1]) + 1)
    version = ".".join(parts)
    metadata["B1"] = version
    metadata["B2"] = datetime.now(timezone.utc).isoformat()
    return version


def import_language(
    csv_path: Path,
    workbook_path: Path,
    language: str,
    *,
    check_only: bool = False,
) -> tuple[int, str | None]:
    translations = load_translation_csv(csv_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Translations"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]

    if "Original" not in headers:
        raise ValueError("dictionary.xlsx has no Original column")
    if language not in headers:
        raise ValueError(f"dictionary.xlsx has no {language!r} column")

    original_column = headers.index("Original") + 1
    language_column = headers.index(language) + 1
    workbook_rows: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        original = sheet.cell(row_number, original_column).value
        if original:
            key = str(original)
            if key in workbook_rows:
                raise ValueError(f"duplicate Original value in workbook row {row_number}")
            workbook_rows[key] = row_number

    csv_keys = set(translations)
    workbook_keys = set(workbook_rows)
    missing_in_workbook = sorted(csv_keys - workbook_keys)
    missing_in_csv = sorted(workbook_keys - csv_keys)
    if missing_in_workbook or missing_in_csv:
        details = []
        if missing_in_workbook:
            details.append(f"{len(missing_in_workbook)} CSV strings missing in workbook")
        if missing_in_csv:
            details.append(f"{len(missing_in_csv)} workbook strings missing in CSV")
        raise ValueError("; ".join(details))

    updates = 0
    for original, translation in translations.items():
        row_number = workbook_rows[original]
        cell = sheet.cell(row_number, language_column)
        if str(cell.value or "") != translation:
            cell.value = translation
            updates += 1

    if check_only or updates == 0:
        workbook.close()
        return updates, None

    version = bump_workbook_version(workbook)
    descriptor, temp_name = tempfile.mkstemp(
        prefix="dictionary-", suffix=".xlsx", dir=workbook_path.parent
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        workbook.close()
        try:
            os.replace(temp_path, workbook_path)
        except PermissionError:
            # Some Windows/OneDrive setups reject replacing an existing XLSX
            # even when it is not open. Copying the completed temporary file
            # still keeps the original intact until serialization succeeds.
            shutil.copyfile(temp_path, workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return updates, version


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("language", help="Language column, for example es419")
    parser.add_argument(
        "--csv",
        type=Path,
        help="Translation CSV (default: translations/<language>.csv)",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Master dictionary workbook",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Report pending updates without writing the workbook",
    )
    args = parser.parse_args()

    csv_path = args.csv or ROOT / "translations" / f"{args.language}.csv"
    try:
        updates, version = import_language(
            csv_path.resolve(),
            args.workbook.resolve(),
            args.language,
            check_only=args.check,
        )
    except (OSError, ValueError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1

    action = "pending" if args.check else "applied"
    print(f"{args.language}: {updates} update(s) {action}")
    if version:
        print(f"dictionary version: {version}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
