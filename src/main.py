"""DBI Translation Pipeline.

Usage:
    python -m src.main sync        — sync ua.csv into dictionary.xlsx
    python -m src.main translate   — translate missing cells via AI
    python -m src.main validate    — validate all translations
    python -m src.main export      — export per-language CSVs
    python -m src.main build       — build .bin files from CSVs
    python -m src.main all         — run full pipeline
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import shutil
import sys
import time
from pathlib import Path

# Force stdout to UTF-8 on Windows
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

from src.core.text_utils import tokenize, detokenize, normalize_tokens_out, visual_length, normalize_fullwidth
from src.core.validator import validate
from src.core.ai_client import translate_batch, refine, init_session, init_session_shadok, translate_shadok_block

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DICT_PATH = DATA_DIR / "dictionary.xlsx"
UA_CSV = DATA_DIR / "ua.csv"
LANG_JSON = DATA_DIR / "languages.json"
OUTPUT_DIR = ROOT / "output"
TRANSLATIONS_DIR = ROOT / "translations"
DIST_DIR = ROOT / "dist"
BUILD_SCRIPT = ROOT / "scripts" / "build_translation_bin.py"

SHEET_NAME = "Translations"
META_SHEET = "Metadata"

BLOCK_JSON = DATA_DIR / "blocks.json"
SHADOK_JSON = DATA_DIR / "shadok.json"

def find_dbi_version_row(ws, col_map) -> int:
    """Find the row that likely contains the pure 3-4 digit version number."""
    import re as _re
    for row in range(2, min(50, ws.max_row + 1)):
        val = str(ws.cell(row, col_map["Original"]).value or "").strip()
        if _re.match(r"^\d{3,4}$", val):
            return row
    return 9  # fallback


def load_shadok_config() -> dict | None:
    """Load shadok.json config. Returns None if not found."""
    if not SHADOK_JSON.exists():
        return None
    with SHADOK_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def get_nro_version() -> str | None:
    """Extract DBI version from patched NRO filename (e.g. DBI.892.ru_patched.nro -> '892').
    Returns the highest version number found from *_patched.nro files."""
    import re as _re
    versions = []
    for nro_file in ROOT.glob("DBI.*.nro"):
        if "debug" in nro_file.name.lower():
            continue
        match = _re.search(r'DBI\.(\d+)\.', nro_file.name)
        if match and nro_file.name.endswith("_patched.nro"):
            versions.append(int(match.group(1)))

    if versions:
        return str(max(versions))
    return None


def get_patched_nro_path() -> Path | None:
    """Get the path to the latest patched NRO file."""
    import re as _re
    nro_files = [f for f in ROOT.glob("DBI.*.nro") if f.name.endswith("_patched.nro") and "debug" not in f.name.lower()]
    if not nro_files:
        return None

    def extract_version(nro_path):
        match = _re.search(r'DBI\.(\d+)\.', nro_path.name)
        return int(match.group(1)) if match else 0

    return max(nro_files, key=extract_version)

# ── helpers ──────────────────────────────────────────────────────────

def load_languages() -> dict[str, str]:
    with LANG_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_blocks() -> dict[str, list[str]]:
    if not BLOCK_JSON.exists():
        return {}
    with BLOCK_JSON.open("r", encoding="utf-8") as f:
        return json.load(f)


def open_or_create_workbook() -> openpyxl.Workbook:
    if DICT_PATH.exists():
        return openpyxl.load_workbook(DICT_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    meta = wb.create_sheet(META_SHEET)
    meta["A1"] = "version"
    meta["B1"] = "0.0.0"
    meta["A2"] = "updated"
    meta["B2"] = ""
    return wb


def get_version(wb: openpyxl.Workbook) -> str:
    meta = wb[META_SHEET]
    return str(meta["B1"].value or "0.0.0")


def bump_version(wb: openpyxl.Workbook) -> str:
    meta = wb[META_SHEET]
    ver = str(meta["B1"].value or "0.0.0")
    parts = ver.split(".")
    parts[-1] = str(int(parts[-1]) + 1)
    new_ver = ".".join(parts)
    meta["B1"] = new_ver
    from datetime import datetime, timezone
    meta["B2"] = datetime.now(timezone.utc).isoformat()
    return new_ver


def save_workbook(wb: openpyxl.Workbook) -> None:
    DICT_PATH.parent.mkdir(parents=True, exist_ok=True)
    max_retries = 10
    for attempt in range(max_retries):
        try:
            wb.save(DICT_PATH)
            return
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n  [!] File '{DICT_PATH.name}' is locked (likely open in Excel).")
                print(f"      PLEASE CLOSE IT! Retrying in 3s... ({attempt + 1}/{max_retries})")
                time.sleep(3)
            else:
                print(f"\n  [CRITICAL] Could not save '{DICT_PATH.name}' after {max_retries} attempts.")
                raise


def sanitize_string(s: str | None) -> str:
    """Remove illegal control characters that openpyxl cannot handle."""
    if s is None:
        return ""
    s = str(s)
    # Remove control characters except \n, \r, \t
    return re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', s)


# ── sync ─────────────────────────────────────────────────────────────

def cmd_sync() -> None:
    """Read ua.csv, tokenize, add missing rows to dictionary.xlsx."""
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    # Ensure header row
    expected_cols = ["Original"] + list(langs.keys())
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        for ci, col in enumerate(expected_cols, 1):
            ws.cell(1, ci, col)

    # Read current header to build col map
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Remove columns not in languages.json (except 'Original')
    valid_cols = {"Original"} | set(langs.keys())
    cols_to_remove = []
    for ci, col_name in enumerate(header):
        if col_name and col_name not in valid_cols:
            cols_to_remove.append((ci + 1, col_name))  # 1-indexed
    
    if cols_to_remove:
        # Delete in reverse order to preserve indices
        for col_idx, col_name in sorted(cols_to_remove, reverse=True):
            ws.delete_cols(col_idx)
            print(f"  Removed column '{col_name}' (not in languages.json)")
        # Re-read header after deletion
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Add missing language columns
    for lang_code in langs:
        if lang_code not in header:
            idx = len(header) + 1
            ws.cell(1, idx, lang_code)
            header.append(lang_code)

    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # Collect existing originals
    existing: set[str] = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val:
            existing.add(val)
            
    # CRITICAL: Since Shadok translations replace 'Original' with their satellite text,
    # the original text from ua.csv will falsely appear as missing.
    # We must explicitly track original Shadok texts and match them dynamically.
    shadok_origs_stripped = set()
    shadok_config = load_shadok_config()
    if shadok_config:
        for item in shadok_config.get("mapping", []):
            if item["new"] in existing:
                # Keep track of the original Shadok texts (stripped)
                shadok_origs_stripped.add(item["orig"].strip())

    # Read ua.csv and insert missing
    added = 0
    with UA_CSV.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for rec in reader:
            ru_raw = rec.get("original", "") or ""
            if not ru_raw:
                continue
            ru_tok = tokenize(ru_raw)
            if ru_tok in existing:
                continue
            
            # Formatting-agnostic check for Shadok strings (which often have leading spaces)
            if ru_tok.strip() in shadok_origs_stripped:
                continue

            next_row = ws.max_row + 1
            ws.cell(next_row, col_map["Original"], ru_tok)
            existing.add(ru_tok)
            added += 1

    # Update DBI version from patched NRO filename
    patched_nro = get_patched_nro_path()
    if patched_nro:
        nro_ver = get_nro_version()
        print(f"  Using patched NRO: {patched_nro.name} (version {nro_ver})")
        version_row = find_dbi_version_row(ws, col_map)
        current_ver = ws.cell(version_row, col_map["Original"]).value
        if str(current_ver) != nro_ver:
            print(f"  Updating DBI version at row {version_row}: {current_ver} -> {nro_ver}")
            # Write version to Original and ALL language columns
            for col_idx in col_map.values():
                ws.cell(version_row, col_idx, nro_ver)
        else:
            print(f"  DBI version already current: {nro_ver}")
    else:
        print("  WARNING: No DBI.*_patched.nro file found, version not updated.")

    ver = bump_version(wb)
    save_workbook(wb)
    print(f"Sync done. Added: {added}, Version: {ver}")


# ── translate ────────────────────────────────────────────────────────

def wrap_text(text: str, max_chars: int, lang_code: str) -> list[str]:
    """Wrap text into lines, with special handling for CJK width."""
    # Heuristic for wider CJK characters as requested by user
    effective_max = max_chars
    # Japanese (jp), Korean (kr), Chinese (zh/zhcn/zhtw)
    if lang_code.lower() in ["zhcn", "zhtw", "jp", "kr", "zh"]:
        # "1.5 times smaller"
        effective_max = int(max_chars / 1.5)
        
    import textwrap
    # replace_whitespace=True converts all tabs/newlines into spaces before wrapping
    lines = textwrap.wrap(text, width=effective_max, break_long_words=True, replace_whitespace=True)
    return lines

def cmd_translate() -> None:
    """Find empty cells and translate via AI with continuous chat."""
    import re
    MAX_REFINE_ATTEMPTS = 3

    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    force_all = "--force" in sys.argv or "-f" in sys.argv

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    lang_codes = [lc for lc in langs if lc in col_map and lc != "ru"]

    # ── Phase -1: CLEANUP DUPLICATE ROWS ───────────────────────────────
    # Rows are considered duplicates if their "Original" column is exactly the same
    best_rows = {}
    rows_to_delete = []
    
    for row in range(2, ws.max_row + 1):
        original_val = str(ws.cell(row, col_map["Original"]).value or "")
        key = original_val  # Use exact string — spaces are significant
        if not key.strip():
            continue
            
        # Count translated languages
        non_empty = 0
        for lc in lang_codes:
            if str(ws.cell(row, col_map[lc]).value or "").strip():
                non_empty += 1
                
        if key in best_rows:
            best_idx, best_count = best_rows[key]
            if non_empty > best_count:
                # The current row is better (more translations)
                rows_to_delete.append(best_idx)
                best_rows[key] = (row, non_empty)
            else:
                # The previous row was better or equal, so we delete the current row
                rows_to_delete.append(row)
        else:
            best_rows[key] = (row, non_empty)
            
    if rows_to_delete:
        # Sort in reverse to delete from bottom to top so indices don't shift
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        print(f"  [CLEANUP] Deleted {len(rows_to_delete)} exact duplicate rows.")
        save_workbook(wb)

    # ── Phase 0: SHADOK BLOCK ────────────────────────────────────────
    # TEMPORARILY DISABLED - will fix later
    shadok_row_set = set()  # rows to exclude from regular translation
    shadok_config = {}
    shadok_all_strings = set()

    # shadok_config = load_shadok_config()
    # shadok_row_set = set()  # rows to exclude from regular translation
    #
    # if shadok_config:
    #     ... (all shadok code commented out)


    # ── Scan: count rows that need translation (excluding shadoks) ───
    rows_to_translate = []

    # Safety net: collect all known Shadok strings to ensure no duplicates slip into the general loop
    # TEMPORARILY DISABLED
    # shadok_all_strings = set()
    # for item in shadok_config.get("mapping", []):
    #     shadok_all_strings.add(item["orig"].strip())
    #     shadok_all_strings.add(item["new"].strip())

    for row in range(2, ws.max_row + 1):
        if row in shadok_row_set:
            continue  # skip strictly mapped shadok rows
            
        original = ws.cell(row, col_map["Original"]).value
        if not original or not str(original).strip():
            continue

        missing = []
        for lc in lang_codes:
            cell_val = ws.cell(row, col_map[lc]).value
            if not cell_val or not str(cell_val).strip():
                missing.append(lc)
            else:
                ok, msg = validate(str(original), str(cell_val), lc)
                if not ok and "English preservation" not in msg:
                    print(f"  [Row {row}][{lc}] Invalid existing translation, scheduling re-translation: {msg}")
                    ws.cell(row, col_map[lc], "")
                    missing.append(lc)
        if missing:
            rows_to_translate.append((row, original, missing))

    total_rows = len(rows_to_translate)
    total_all = ws.max_row - 1  # excluding header

    if total_rows == 0:
        print(f"  Nothing to translate. All {total_all} rows are complete.")
        return

    # ── Phase 1: INIT ────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  DBI TRANSLATOR")
    print("=" * 60)
    print(f"  Languages : {len(lang_codes)} ({', '.join(lang_codes)})")
    print(f"  Total rows: {total_all}")
    print(f"  To translate: {total_rows} (excl. {len(shadok_row_set)} shadok rows)")
    print("-" * 60)

    init_session()

    print("-" * 60)
    print(f"  Starting translation...")
    print("=" * 60)
    print()


    # ── Phase 2: TRANSLATE ───────────────────────────────────────────
    total_translated = 0
    total_failed = 0

    for idx, (row, original, missing) in enumerate(rows_to_translate, 1):
        print(f"  [Row {row} | {idx}/{total_rows}] {original[:50]}  -> {len(missing)} langs")

        # Skip AI translation if string has no Cyrillic characters — copy as-is
        cyrillic_count = len(re.findall(r'[а-яА-ЯёЁіІїЇєЄґҐ]', original))
        if cyrillic_count < 2:
            for lc in missing:
                # Auto-fix: replace 'ru' language code with target language code
                translation = original
                if re.search(r'\bru\b', original, re.IGNORECASE):
                    translation = re.sub(r'\bru\b', lc, original, flags=re.IGNORECASE)
                ws.cell(row, col_map[lc], translation)
                total_translated += 1
            save_workbook(wb)
            print(f"    [Row {row} | {idx}/{total_rows}] Cyrillic count {cyrillic_count} <= 3 — copied as-is.")
            continue

        try:
            results = translate_batch(original, missing, row_id=row)
        except Exception as e:
            print(f"  [Row {row} | {idx}/{total_rows}] ERROR: {e}")
            total_failed += len(missing)
            time.sleep(2)
            continue

        # Normalize tokens and full-width chars in AI output
        for lc in list(results.keys()):
            results[lc] = normalize_fullwidth(normalize_tokens_out(results.get(lc, "")))

        # ── Validate + Refine loop ───────────────────────────────────
        for attempt in range(MAX_REFINE_ATTEMPTS):
            errors = []
            for lc in missing:
                translation = results.get(lc, "")
                if not translation:
                    errors.append((lc, "Translation is empty"))
                    continue
                ok, msg = validate(original, translation, lc)
                if not ok:
                    errors.append((lc, msg))

            if not errors:
                break

            if attempt < MAX_REFINE_ATTEMPTS - 1:
                error_lines = "\n".join(f"- {lc}: {msg}" for lc, msg in errors)
                correction = (
                    f"The following translations have errors:\n"
                    f"{error_lines}\n\n"
                    f"Please fix them. Source text: \"{original}\""
                )
                print(f"    [Row {row} | {idx}/{total_rows}] Refine #{attempt + 2}: {len(errors)} errors")
                try:
                    refined = refine(correction, missing, row_id=row)
                    for lc in list(refined.keys()):
                        refined[lc] = normalize_fullwidth(normalize_tokens_out(refined.get(lc, "")))
                    results.update(refined)
                except Exception as e:
                    print(f"    [Row {row} | {idx}/{total_rows}] Refine error: {e}")
                    break
                time.sleep(0.3)

        # ── Write results ────────────────────────────────────────────
        row_ok = 0
        row_fail = 0
        failed_langs = []
        for lc in missing:
            translation = results.get(lc, "")
            if not translation or not translation.strip():
                failed_langs.append(lc)
                continue

            # Auto-fix: replace 'ru' language code with target language code
            if re.search(r'\bru\b', original, re.IGNORECASE):
                translation = re.sub(r'\bru\b', lc, translation, flags=re.IGNORECASE)

            ok, msg = validate(original, translation, lc)
            if ok:
                ws.cell(row, col_map[lc], translation)
                total_translated += 1
                row_ok += 1
            elif "English preservation" in msg:
                ws.cell(row, col_map[lc], original)
                total_translated += 1
                row_ok += 1
            else:
                failed_langs.append(lc)

        # ── Retry failed languages with error context ─────────────────
        MAX_RETRY_ROUNDS = 3
        for retry_round in range(MAX_RETRY_ROUNDS):
            if not failed_langs:
                break

            # Build error context for the AI
            error_details = []
            for lc in failed_langs:
                translation = results.get(lc, "")
                _, msg = validate(original, translation, lc) if translation else (False, "empty")
                error_details.append(f"- {lc}: {msg}")

            error_context = (
                f"Retry round {retry_round + 1}. The following translations for "
                f"\"{original}\" have validation errors:\n"
                + "\n".join(error_details)
                + "\n\nPlease fix these issues. "
                f"Source text: \"{original}\""
            )

            print(f"    [Row {row}] Retry {retry_round + 1}/{MAX_RETRY_ROUNDS}: {len(failed_langs)} langs ({', '.join(failed_langs)})")

            try:
                retry_results = refine(error_context, failed_langs, row_id=row)
                still_failed = []
                for lc in failed_langs:
                    translation = normalize_fullwidth(normalize_tokens_out(retry_results.get(lc, "")))
                    if not translation or not translation.strip():
                        still_failed.append(lc)
                        continue

                    # Auto-fix: replace 'ru' language code with target language code
                    import re
                    if re.search(r'\bru\b', original, re.IGNORECASE):
                        translation = re.sub(r'\bru\b', lc, translation, flags=re.IGNORECASE)

                    ok, msg = validate(original, translation, lc)
                    if ok:
                        ws.cell(row, col_map[lc], translation)
                        total_translated += 1
                        row_ok += 1
                    else:
                        results[lc] = translation  # update for next round's error context
                        still_failed.append(lc)
                failed_langs = still_failed
            except Exception as e:
                print(f"    [Row {row}] Retry error: {e}")
                break
            time.sleep(0.3)

        # Mark remaining failed langs
        for lc in failed_langs:
            print(f"    [Row {row}][{lc}] SKIPPED (invalid after {MAX_RETRY_ROUNDS} retries)")
            total_failed += 1
            row_fail += 1

        save_workbook(wb)

        status = "OK" if row_fail == 0 else f"OK:{row_ok} FAIL:{row_fail}"
        print(f"    [Row {row} | {idx}/{total_rows}] Saved. {status}")

        time.sleep(0.5)

    # ── Phase 3: SUMMARY ─────────────────────────────────────────────
    print()
    print("=" * 60)
    ver = bump_version(wb)
    save_workbook(wb)
    print(f"  DONE! Translated: {total_translated}, Failed: {total_failed}")
    print(f"  Version: {ver}")
    print("=" * 60)




# ── validate ─────────────────────────────────────────────────────────

def cmd_validate() -> None:
    """Validate all translations in the dictionary using advanced Validator."""
    from src.core.validator import Validator
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # Pass BLOCK_JSON to validator for regex-aware checks
    validator = Validator(str(BLOCK_JSON) if BLOCK_JSON.exists() else None)

    # Build shadok exclusion set (use cached rows if available)
    shadok_row_set = set()
    shadok_config = load_shadok_config()
    if shadok_config:
        if "rows" in shadok_config and shadok_config["rows"]:
            shadok_row_set = {r for r in shadok_config["rows"] if r is not None}
        else:
            excel_lookup = {}
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row, col_map["Original"]).value
                if val:
                    excel_lookup[str(val).strip()] = row

            for item in shadok_config.get("mapping", []):
                orig_text = item["orig"].strip()
                new_text = item["new"].strip()
                r = excel_lookup.get(orig_text) or excel_lookup.get(new_text)
                if r:
                    shadok_row_set.add(r)
        if shadok_row_set:
            print(f"  Excluding {len(shadok_row_set)} shadok rows from validation.")

    errors = 0
    checked = 0

    print(f"Starting validation for {len([l for l in langs if l != 'ru'])} languages...")

    for row in range(2, ws.max_row + 1):
        if row in shadok_row_set:
            continue  # skip shadok rows
        original = ws.cell(row, col_map["Original"]).value
        if not original:
            continue
        original_str = str(original)
        
        for lc in langs:
            if lc not in col_map or lc == "ru":
                continue
            translation = ws.cell(row, col_map[lc]).value
            if not translation:
                continue
            
            checked += 1
            row_errors = validator.validate_row(original_str, str(translation), lc)
            if row_errors:
                errors += len(row_errors)
                print(f"  [Row {row}][{lc}] Error(s):")
                for err in row_errors:
                    print(f"    - {err}")

    print(f"\n--- Translation checks: {checked} checked, {errors} issues ---")

    # Phase 2: Exact block validation (blocks.json vs Original column)
    if validator.blocks:
        print("\nRunning exact block validation (blocks.json)...")
        originals = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col_map["Original"]).value
            if val:
                originals[row] = str(val)

        block_errors = validator.validate_blocks(originals)
        if block_errors:
            errors += len(block_errors)
            for err in block_errors:
                print(f"  {err}")
            print(f"--- Block regex checks: {len(block_errors)} issues ---")
        else:
            print("--- Block regex checks: all OK ---")

    print(f"\nValidation complete. Total issues: {errors}")


# ── export ───────────────────────────────────────────────────────────

def cmd_export() -> None:
    """Export per-language CSV files (original, translation) for the bin builder."""
    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    TRANSLATIONS_DIR.mkdir(parents=True, exist_ok=True)

    missing_total = 0
    
    for lc in langs:
        if lc not in col_map or lc == "ru":
            continue
        csv_path = TRANSLATIONS_DIR / f"{lc}.csv"
        count = 0
        missing_count = 0
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["original", "translation"])
            for row in range(2, ws.max_row + 1):
                original = ws.cell(row, col_map["Original"]).value
                if not original:
                    continue
                    
                translation = ws.cell(row, col_map[lc]).value
                
                if not translation or not str(translation).strip():
                    english_fallback = ws.cell(row, col_map["en"]).value if "en" in col_map else None
                    translation = english_fallback if english_fallback and str(english_fallback).strip() else original
                    missing_count += 1
                    missing_total += 1
                    print(f"  [WARNING] Row {row} missing translation for '{lc}'. Fallback to: {translation[:30] + '...' if len(str(translation)) > 30 else translation}")

                writer.writerow([detokenize(str(original)), detokenize(str(translation))])
                count += 1
                
        print(f"  {lc}.csv: {count} entries" + (f" ({missing_count} missing translations filled with fallback)" if missing_count else ""))

    if missing_total > 0:
        print(f"\n[ALERT] Export finished with {missing_total} missing translations!")
        print("Run `python -m src.main translate` to translate missing lines.")
    else:
        print("\nExport done.")

    print("\n[BUILD] Auto-building binaries...")
    cmd_build()


def cmd_sync() -> None:
    """Sync Excel dictionary with translations/ua.csv and ru.csv."""
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # 1. Collect all current keys from Excel
    excel_keys = {}  # original_text -> list of row_indices
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val:
            v_str = str(val)
            if v_str not in excel_keys:
                excel_keys[v_str] = []
            excel_keys[v_str].append(row)

    # 2. Collect all valid keys from source CSVs
    source_data = {} # key -> tokenized_text

    # Read from DATA_DIR (Source of truth), NOT translations/
    for csv_file in [DATA_DIR / "ua.csv", DATA_DIR / "ru.csv"]:
        path = Path(csv_file)
        if not path.exists(): continue

        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 1:
                    orig_tok = tokenize(row[0])
                    if not orig_tok: continue
                    if orig_tok not in source_data:
                        trans = tokenize(row[1]) if len(row) > 1 else orig_tok
                        source_data[orig_tok] = trans

    # 3. Synchronize: Add missing, update existing
    # NOTE: We do NOT delete rows that are not in source CSVs anymore
    # This preserves manually added rows in the dictionary
    added = 0

    for orig_tok, trans_tok in source_data.items():
        if orig_tok not in excel_keys:
            # Add new row
            new_row = ws.max_row + 1
            ws.cell(new_row, col_map["Original"], orig_tok)
            if "ru" in col_map: ws.cell(new_row, col_map["ru"], trans_tok)
            if "ua" in col_map: ws.cell(new_row, col_map["ua"], trans_tok)
            added += 1
            excel_keys[orig_tok] = [new_row]

    # 4. Remove duplicates only (keep first occurrence)
    duplicates_removed = 0
    for orig, row_indices in excel_keys.items():
        if len(row_indices) > 1:
            print(f"  [SYNC] Removing {len(row_indices)-1} duplicate(s) for: {repr(orig[:30])}...")
            # Delete all duplicates except the first one
            for row_idx in sorted(row_indices[1:], reverse=True):
                ws.delete_rows(row_idx)
                duplicates_removed += 1

    # Bump version or just update status
    ver = get_version(wb)
    print(f"Sync done. Added: {added}, Duplicates removed: {duplicates_removed}, Version: {ver}")
    save_workbook(wb)


# ── build ────────────────────────────────────────────────────────────

def cmd_build() -> None:
    """Run build_translation_bin.py for each exported CSV."""
    langs = load_languages()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for lc in langs:
        if lc == "ru":
            continue
        csv_path = TRANSLATIONS_DIR / f"{lc}.csv"
        if not csv_path.exists():
            print(f"  Skip {lc}: no CSV")
            continue
        bin_path = OUTPUT_DIR / f"translation_{lc}.bin"
        cmd = [sys.executable, str(BUILD_SCRIPT), str(csv_path), "-o", str(bin_path)]
        print(f"  Building {bin_path.name}...")
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  [!] Error building {lc}: {result.stderr}")
        else:
            print(f"  OK: {bin_path.name}")

    print("Build done.")


def cmd_dist() -> None:
    """Organize NRO and translation bins into per-language folders in 'dist'."""
    langs = load_languages()

    if DIST_DIR.exists():
        for attempt in range(5):
            try:
                shutil.rmtree(DIST_DIR)
                break
            except PermissionError:
                if attempt < 4:
                    print(f"  [!] dist/ is locked. Retrying in 2s... ({attempt + 1}/5)")
                    time.sleep(2)
                else:
                    print("  [CRITICAL] Could not clean dist/ — close Explorer or other programs using it.")
                    raise
    DIST_DIR.mkdir(parents=True)

    # Find the latest patched DBI NRO file
    source_nro = get_patched_nro_path()
    if not source_nro:
        print("  [ERROR] No DBI.*_patched.nro file found in root!")
        return

    nro_ver = get_nro_version()
    print(f"  Using patched NRO: {source_nro.name} (version {nro_ver})")

    for lc in langs:
        if lc == "ru": continue

        bin_path = OUTPUT_DIR / f"translation_{lc}.bin"
        if not bin_path.exists():
            # Try to build it if missing? No, user usually runs build before dist.
            continue

        lang_dist = DIST_DIR / lc
        lang_dist.mkdir(parents=True, exist_ok=True)

        # Copy and rename NRO to DBI.nro as requested by user's example
        shutil.copy2(source_nro, lang_dist / "DBI.nro")
        # Copy and rename BIN to translation.bin
        shutil.copy2(bin_path, lang_dist / "translation.bin")

        print(f"  [OK] {lc}: DBI.nro + translation.bin")

    print(f"\nOrganization in 'dist' folder complete using {source_nro.name}")


def cmd_align() -> None:
    """Align colons in blocks by longest line per language per block."""

    blocks = load_blocks()
    if not blocks:
        print("No blocks defined in blocks.json.")
        return

    langs = load_languages()
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    # Cache all Original values for faster matching
    originals = {}  # row -> original_str
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_map["Original"]).value
        if val:
            originals[row] = str(val)

    # Build grouped_data[block_id] = list of matched row indices
    grouped_data = {}  # block_id -> [row_idx, ...]

    for bid, patterns in blocks.items():
        matched_rows = []
        for pattern in patterns:
            found = False
            for row, orig_val in originals.items():
                if orig_val == pattern:
                    matched_rows.append(row)
                    found = True
                    break

            if not found:
                print(f"  [WARN] String not found for block {bid}: {pattern[:50]}...")

        if matched_rows:
            grouped_data[bid] = matched_rows

    if not grouped_data:
        print("No rows found matching blocks.json patterns.")
        return

    lang_cols = {lc: col_map[lc] for lc in langs if lc in col_map and lc != "ru"}

    affected_count = 0

    # Process each block
    for bid, matched_rows in grouped_data.items():
        print(f"  Aligning block: {bid:<25} (Rows: {len(matched_rows)})")

        # Special handling for NSP_INSTALL_ANSWERS: pad to exactly 4 chars
        if bid == "NSP_INSTALL_ANSWERS":
            for lc, col_idx in lang_cols.items():
                for row_idx in matched_rows:
                    val = ws.cell(row_idx, col_idx).value
                    if not val:
                        continue
                    val_str = str(val)
                    stripped = val_str.strip()
                    char_count = len(stripped)

                    if char_count >= 4:
                        new_val = stripped[:4]
                    elif char_count == 3:
                        new_val = stripped + " "
                    elif char_count == 2:
                        new_val = " " + stripped + " "
                    elif char_count == 1:
                        new_val = " " + stripped + "  "
                    else:
                        new_val = "    "

                    if new_val != val_str:
                        ws.cell(row_idx, col_idx, new_val)
                        affected_count += 1
            continue

        # Special handling for TITLE_INFO block with double-colon strings
        is_title_info = (bid == "TITLE_INFO")

        # For each language, find max length and align
        for lc, col_idx in lang_cols.items():
            # Find max prefix length for this language in this block
            max_len = 0
            for row_idx in matched_rows:
                val = ws.cell(row_idx, col_idx).value
                if not val:
                    continue
                val_str = str(val).strip()

                if ":" in val_str:
                    # For TITLE_INFO: only consider first colon for alignment
                    if is_title_info:
                        prefix = val_str.split(":", 1)[0]
                    else:
                        prefix = val_str.split(":", 1)[0]
                    clean_prefix = prefix.rstrip()
                    max_len = max(max_len, len(clean_prefix))

            if max_len == 0:
                continue

            # Align all rows in this block for this language
            target_len = max_len + 1  # +1 for at least one space before colon

            for row_idx in matched_rows:
                val = ws.cell(row_idx, col_idx).value
                if not val:
                    continue
                val_str = str(val).strip()

                if ":" in val_str:
                    prefix, suffix = val_str.split(":", 1)
                    clean_prefix = prefix.rstrip()
                    current_len = len(clean_prefix)

                    padding = target_len - current_len
                    if padding < 1:
                        padding = 1

                    new_val = clean_prefix + (' ' * padding) + ':' + suffix

                    if new_val != val_str:
                        ws.cell(row_idx, col_idx, new_val)
                        affected_count += 1

    if affected_count > 0:
        ver = bump_version(wb)
        save_workbook(wb)
        print(f"\nAlignment done. Adjusted {affected_count} cells. Version: {ver}")
    else:
        print("\nAlignment done. No changes needed.")




# ── clear ────────────────────────────────────────────────────────────

def cmd_clear(lang_code: str) -> None:
    """Clear all translations for a specific language in the dictionary."""
    if not DICT_PATH.exists():
        print("No dictionary found.")
        return

    wb = openpyxl.load_workbook(DICT_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet {SHEET_NAME} not found.")
        return

    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}

    if lang_code not in col_map:
        print(f"Language '{lang_code}' not found in dictionary columns.")
        return

    col_idx = col_map[lang_code]
    cleared = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, col_idx).value:
            ws.cell(row, col_idx).value = None
            cleared += 1

    ver = bump_version(wb)
    save_workbook(wb)
    print(f"Cleared {cleared} entries for '{lang_code}'. Version: {ver}")


# ── main ─────────────────────────────────────────────────────────────

def cmd_deploy() -> None:
    """Commit, push and create a GitHub release with assets."""

    # 1. Get versions & Check completeness
    try:
        patched_nro = get_patched_nro_path()
        if not patched_nro:
            print("  [ERROR] No DBI.*_patched.nro file found!")
            return

        dbi_ver = get_nro_version()
        print(f"  Using patched NRO: {patched_nro.name} (version {dbi_ver})")

        wb = open_or_create_workbook()
        ws = wb[SHEET_NAME]
        patcher_ver = get_version(wb)

        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {h: i + 1 for i, h in enumerate(header) if h}
        langs = load_languages()
        lang_codes = [lc for lc in langs if lc in col_map and lc != "ru"]

        print("  [CHECK] Verifying translation completeness...")
        missing_count = 0
        for row in range(2, ws.max_row + 1):
            if not ws.cell(row, col_map["Original"]).value: continue
            for lc in lang_codes:
                val = ws.cell(row, col_map[lc]).value
                if not val or not str(val).strip():
                    missing_count += 1

        if missing_count > 0:
            print(f"  [ERROR] Cannot deploy: Found {missing_count} missing translations!")
            print("  Please run 'python -m src.main translate' first.")
            return

    except Exception as e:
        print(f"  [ERROR] Preparation failed: {e}")
        return

    # 2. Copy files to target directories (always execute)
    print("  [COPY] Copying files to target directories...")
    try:
        # Find the patched NRO
        patched_nro = get_patched_nro_path()

        if not patched_nro:
            print(f"  [WARN] No patched NRO found for version {dbi_ver}")
        else:
            print(f"  [COPY] Using: {patched_nro.name}")
            # Copy to D:\git\dev\_kefir\kefir\switch\DBI\DBI.nro
            kefir_dir = Path("D:/git/dev/_kefir/kefir/switch/DBI")
            kefir_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(patched_nro, kefir_dir / "DBI.nro")
            print(f"  [COPY] {patched_nro.name} -> {kefir_dir / 'DBI.nro'}")

            # Copy translation_en.bin to D:\git\dev\_kefir\kefir\switch\DBI\translation.bin
            en_bin = OUTPUT_DIR / "translation_en.bin"
            if en_bin.exists():
                shutil.copy2(en_bin, kefir_dir / "translation.bin")
                print(f"  [COPY] translation_en.bin -> {kefir_dir / 'translation.bin'}")
            else:
                print(f"  [WARN] translation_en.bin not found")

            # Copy translation_ua.bin to E:\Switch\addons\switch\DBI\translation.bin
            ua_bin = OUTPUT_DIR / "translation_ua.bin"
            switch_dir = Path("E:/Switch/addons/switch/DBI")
            if ua_bin.exists():
                switch_dir.mkdir(parents=True, exist_ok=True)
                shutil.copy2(ua_bin, switch_dir / "translation.bin")
                print(f"  [COPY] translation_ua.bin -> {switch_dir / 'translation.bin'}")
            else:
                print(f"  [WARN] translation_ua.bin not found")

        print("  [COPY] File copying completed!")
    except Exception as e:
        print(f"  [ERROR] File copying failed: {e}")

    # 3. Git operations
    print("  [GIT] Staging changes and pushing...")
    try:
        subprocess.run(["git", "add", "."], check=True)
        # Check if there are changes to commit
        status = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, encoding="utf-8").stdout
        if status.strip():
            subprocess.run(["git", "commit", "-m", f"chore: deploy DBI {dbi_ver} localization (v{patcher_ver})"], check=True)
            subprocess.run(["git", "push", "origin", "master"], check=True)
            print("  [GIT] Changes pushed successfully.")
        else:
            print("  [GIT] No changes to commit.")
    except subprocess.CalledProcessError as e:
        print(f"  [ERROR] Git operation failed: {e}")
        return

    # 4. Prepare release body
    # (template remains same)
    langs_list = """*   **BE** — Belarusian
*   **DE** — German
*   **EN** — English (US)
*   **ENGB** — English (UK)
*   **ES** — Spanish (Spain)
*   **ES419** — Spanish (Latin America)
*   **ET** — Estonian
*   **FR** — French
*   **FRCA** — French (Canada)
*   **IT** — Italian
*   **JP** — Japanese
*   **KK** — Kazakh
*   **KR** — Korean
*   **LT** — Lithuanian
*   **LV** — Latvian
*   **NL** — Dutch
*   **PL** — Polish
*   **PT** — Portuguese (Portugal)
*   **PTBR** — Portuguese (Brazil)
*   **UA** — Ukrainian
*   **ZHCN** — Chinese (Simplified)
*   **ZHTW** — Chinese (Traditional)"""

    release_body = f"""### 🌍 DBI Multilingual Localization (v{patcher_ver})
![GitHub release (tag)](https://img.shields.io/github/downloads/rashevskyv/DBIPatcher/{dbi_ver}/total)

This release provides high-quality translations for **DBI version {dbi_ver}**.

> [!IMPORTANT]
> This translation is **strictly compatible only with the DBI.nro version provided in this release**. Do not use it with other versions of DBI as it may cause UI glitches or crashes.

### 📦 Supported Languages
{langs_list}

***

### 🛠️ Installation Instruction
1. Download **`DBI.nro`** (Patched/Compatible version) from this release.
2. Download the **`translation_XX.bin`** file for your desired language.
3. **Rename** the translation file to exactly `translation.bin`.
4. Place both `DBI.nro` and `translation.bin` into the `/switch/DBI/` folder on your SD card.

### ⚠️ Known Issues
- ~~**Hardcoded Strings**: Some interface elements are hardcoded within the DBI binary and cannot be localized via `translation.bin`. Confirmation prompts may still display in Russian: **Да** (Yes) and **Нет** (No).~~ ✅ Fixed!
- **Shadok Fables**: Satirical text blocks and stories (Shadok fables) remain in their original form.
- ~~**System Language Names**: Names of languages in the DBI settings menu are hardcoded in the binary.~~ ✅ Fixed!
- **Launcher Compatibility**: Translations have been tested exclusively on [Kefir](https://github.com/rashevskyv/kefir). On Kefir, the translation works successfully regardless of whether DBI is launched via [Sphaira](https://github.com/ITotalJustice/sphaira) or [nx-hbmenu](https://github.com/switchbrew/nx-hbmenu/releases/). If you experience issues with translations not applying on other custom firmwares, please refer to [#12](https://github.com/rashevskyv/DBIPatcher/issues/12).

***
*Note: This NRO is a modified version of the [original DBI](https://github.com/rashevskyv/dbi/releases/tag/{dbi_ver}ru) optimized for these translations.*
"""
    
    body_path = Path("scratch/release_body.md")
    body_path.parent.mkdir(exist_ok=True)
    body_path.write_text(release_body, encoding="utf-8")

    # 5. GitHub Release - Create or update
    print(f"  [GH] Checking if release {dbi_ver} exists...")
    check_tag = subprocess.run(["gh", "release", "view", dbi_ver], capture_output=True, text=True, encoding="utf-8")

    assets = [str(a) for a in Path("output").glob("translation_*.bin")]

    # Get the patched NRO for release and rename it for the asset upload
    patched_nro = get_patched_nro_path()
    nro_assets = []
    if patched_nro:
        print(f"  [GH] Preparing NRO asset: {patched_nro.name} -> DBI.nro")
        release_nro = Path("scratch/DBI.nro")
        shutil.copy2(patched_nro, release_nro)
        nro_assets = [str(release_nro)]
    else:
        print(f"  [WARN] No patched NRO found to include in release")

    if check_tag.returncode == 0:
        # Release exists - update assets and notes
        print(f"  [GH] Release {dbi_ver} exists, updating assets and notes...")

        # Detect if NRO changed by comparing local file size with release asset
        nro_changed = False
        if patched_nro and patched_nro.exists():
            local_nro_size = patched_nro.stat().st_size
            asset_info = subprocess.run(
                ["gh", "release", "view", dbi_ver, "--json", "assets", "-q", ".assets[] | select(.name==\"DBI.nro\") | .size"],
                capture_output=True, text=True, encoding="utf-8"
            )
            remote_nro_size = int(asset_info.stdout.strip()) if asset_info.stdout.strip().isdigit() else 0
            if remote_nro_size != local_nro_size:
                nro_changed = True
                print(f"  [GH] NRO changed: remote={remote_nro_size} vs local={local_nro_size}")

        # Add update notice to release body
        from datetime import datetime, timezone, timedelta
        kyiv_tz = timezone(timedelta(hours=3))
        kyiv_time = datetime.now(kyiv_tz).strftime("%Y-%m-%d %H:%M")

        redownload_items = "**translation files**"
        if nro_changed:
            redownload_items = "**DBI.nro** and **translation files**"

        update_notice = f"""> [!WARNING]
> 🔄 **Release updated on {kyiv_time} (Kyiv time).** Please re-download {redownload_items} to get the latest version.
"""
        # Insert update notice after the badge line
        badge_line = f"![GitHub release (tag)](https://img.shields.io/github/downloads/rashevskyv/DBIPatcher/{dbi_ver}/total)"
        release_body = release_body.replace(
            badge_line,
            badge_line + "\n\n" + update_notice
        )
        body_path.write_text(release_body, encoding="utf-8")

        # Update release notes
        try:
            subprocess.run(["gh", "release", "edit", dbi_ver, "--notes-file", str(body_path)], check=True)
            print(f"  [GH] Release notes updated for {dbi_ver}")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] Failed to update release notes: {e}")

        # Upload all assets with --clobber to overwrite existing files
        upload_cmd = ["gh", "release", "upload", dbi_ver, "--clobber"]
        upload_cmd.extend(assets)
        upload_cmd.extend(nro_assets)

        try:
            subprocess.run(upload_cmd, check=True)
            print(f"  [GH] Assets updated successfully in release {dbi_ver}!")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] Failed to update assets: {e}")
    else:
        # Release doesn't exist - create new one
        print(f"  [GH] Creating new release {dbi_ver}...")

        cmd = [
            "gh", "release", "create", dbi_ver,
            "--title", f"DBI {dbi_ver} Localization",
            "--notes-file", str(body_path)
        ]
        cmd.extend(assets)
        cmd.extend(nro_assets)

        try:
            subprocess.run(cmd, check=True)
            print(f"  [GH] Release {dbi_ver} created successfully!")
        except subprocess.CalledProcessError as e:
            print(f"  [ERROR] GitHub release failed: {e}")


def cmd_check() -> None:
    """Check dictionary integrity against source CSV and blocks.json"""
    print("\n" + "="*60 + "\n  STEP: check\n" + "="*60)
    
    ua_path = DATA_DIR / "ua.csv"
    ua_originals = []
    if ua_path.exists():
        with open(ua_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if row:
                    ua_originals.append(tokenize(row[0]))
                    
    wb = open_or_create_workbook()
    ws = wb[SHEET_NAME]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(header) if h}
    
    dict_originals = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col_map["Original"]).value
        if v:
            dict_originals.add(str(v))
            
    mismatches = 0
    for k in ua_originals:
        if k not in dict_originals:
            print(f"  [ERROR] Source key missing or altered in dictionary: {repr(k)}")
            mismatches += 1
            
    blocks = load_blocks()
    if blocks:
        for bid, patterns in blocks.items():
            for pat in patterns:
                if pat not in dict_originals:
                    print(f"  [ERROR] Block {bid} string missing in dictionary: {repr(pat)}")
                    mismatches += 1
                
    if mismatches == 0:
        print("  [OK] Health check passed successfully! All strings match identically.")
    else:
        print(f"\n  [FAIL] Health check failed with {mismatches} issues.")


COMMANDS = {
    "sync": cmd_sync,
    "translate": cmd_translate,
    "validate": cmd_validate,
    "align": cmd_align,
    "export": cmd_export,
    "build": cmd_build,
    "dist": cmd_dist,
    "clear": cmd_clear,
    "deploy": cmd_deploy,
    "check": cmd_check,
    "test": lambda: cmd_test(),
}


def cmd_all() -> None:
    for name in ("sync", "translate", "align", "validate", "export", "build", "dist"):
        print(f"\n{'='*60}\n  STEP: {name}\n{'='*60}")
        COMMANDS[name]()


def cmd_test() -> None:
    """Run all steps except deploy (sync, translate, align, validate, export, build, dist)."""
    for name in ("sync", "translate", "align", "validate", "export", "build", "dist"):
        print(f"\n{'='*60}\n  STEP: {name}\n{'='*60}")
        COMMANDS[name]()
    print(f"\n{'='*60}\n  TEST COMPLETE\n{'='*60}")
    print("All steps completed successfully. Ready for deployment.")
    print("To deploy, run: python -m src.main deploy")


def cmd_help() -> None:
    """Display all available commands with descriptions."""
    print("\n" + "="*60)
    print("  DBI TRANSLATION PIPELINE - Available Commands")
    print("="*60)
    print("\nUsage: python -m src.main <command> [command2 ...] [options]\n")
    print("Commands:")
    print("  sync        - Sync ua.csv into dictionary.xlsx")
    print("  translate   - Translate missing cells via AI")
    print("  validate    - Validate all translations")
    print("  align       - Align colons in blocks by longest line")
    print("  export      - Export per-language CSVs")
    print("  build       - Build .bin files from CSVs")
    print("  dist        - Organize NRO and bins into dist folders")
    print("  clear       - Clear all translations for a language")
    print("                Usage: python -m src.main clear <lang_code>")
    print("  deploy      - Commit, push and create GitHub release")
    print("  check       - Check dictionary integrity")
    print("  test        - Run all steps except deploy")
    print("  all         - Run full pipeline (sync → dist)")
    print("  help        - Show this help message")
    print("\nOptions:")
    print("  -f, --force - Force re-translate all strings")
    print("\nExamples:")
    print("  python -m src.main sync")
    print("  python -m src.main translate -f")
    print("  python -m src.main align build")
    print("  python -m src.main export build dist")
    print("  python -m src.main clear ua")
    print("  python -m src.main all")
    print("="*60 + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="DBI Translation Pipeline", add_help=False)
    parser.add_argument("commands", nargs="*", help="Pipeline steps to run")
    parser.add_argument("-f", "--force", action="store_true", help="Force re-translate all strings")
    parser.add_argument("-h", "--help", action="store_true", help="Show help message")
    args = parser.parse_args()

    if args.help or not args.commands or (len(args.commands) == 1 and args.commands[0] == "help"):
        cmd_help()
        return 0

    # Validate commands
    valid_commands = {*COMMANDS.keys(), "all", "help"}
    for cmd in args.commands:
        if cmd not in valid_commands and cmd not in load_languages():
            print(f"Error: Unknown command '{cmd}'")
            print(f"Valid commands: {', '.join(sorted(valid_commands))}")
            return 1

    # Clear log at every start
    log_path = ROOT / "logs" / "ai_proxy.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    from datetime import datetime, timezone
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"--- SESSION: {' '.join(args.commands)} | {datetime.now(timezone.utc).isoformat()} ---\n")

    # Execute commands sequentially
    for i, cmd in enumerate(args.commands):
        if len(args.commands) > 1:
            print(f"\n{'='*60}\n  STEP {i+1}/{len(args.commands)}: {cmd}\n{'='*60}")

        if cmd == "all":
            cmd_all()
        elif cmd == "clear":
            # For clear command, next argument should be language code
            if i + 1 >= len(args.commands):
                print("Error: 'clear' command requires a language code (e.g., 'ua')")
                return 1
            lang_code = args.commands[i + 1]
            cmd_clear(lang_code)
            # Skip next argument as it was the language code
            args.commands[i + 1] = None
        elif cmd is None:
            # Skip (was consumed as language code for clear)
            continue
        else:
            COMMANDS[cmd]()

    return 0


if __name__ == "__main__":
    sys.exit(main())
